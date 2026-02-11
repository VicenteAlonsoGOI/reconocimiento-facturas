from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self):
        # Colores corporativos (ejemplo: Azul Oscuro y Gris Claro)
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        
        self.zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

    def crear_excel_expediente(self, datos_facturas, output_path):
        """
        Crea un Excel con los datos de múltiples facturas.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas Extraídas"
        
        # Nuevas cabeceras detalladas
        headers = [
            "Factura", "Fecha Factura", "Fecha Cargo", "Base Imponible", "Total", 
            "BI 5%", "IVA 5%", "BI 10%", "IVA 10%", "BI 21%", "IVA 21%", 
            "CHECK TOTAL", "IVA (Desglose)", "Archivo Original"
        ]
        ws.append(headers)
        
        # Estilo para cabecera
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
            
        # Añadir datos
        for idx, factura in enumerate(datos_facturas, 2):
            # Procesar IVA (puede ser dic o string "0%")
            iva_data = factura.get('IVA')
            
            # Valores por defecto
            bi_5 = iva_5 = bi_10 = iva_10 = bi_21 = iva_21 = 0.0
            iva_texto = "0%"

            if isinstance(iva_data, dict):
                # Extraer valores del diccionario estructurado
                bi_5 = iva_data.get('5', {}).get('base', 0.0)
                iva_5 = iva_data.get('5', {}).get('cuota', 0.0)
                bi_10 = iva_data.get('10', {}).get('base', 0.0)
                iva_10 = iva_data.get('10', {}).get('cuota', 0.0)
                bi_21 = iva_data.get('21', {}).get('base', 0.0)
                iva_21 = iva_data.get('21', {}).get('cuota', 0.0)
                
                # Reconstruir texto resumen para la columna informativa
                partes = []
                for k, v in iva_data.items():
                    if v['cuota'] > 0:
                        partes.append(f"{k}%: {v['cuota']}")
                iva_texto = ", ".join(partes) if partes else "0%"
            else:
                iva_texto = str(iva_data)

            # Extraer Base Imponible total (Col D)
            base_imp = factura.get('Base Imponible', 0.0)
            
            row_data = [
                factura.get('Factura'),
                factura.get('Fecha Factura'),
                factura.get('Fecha Cargo'),
                base_imp,
                factura.get('Total'),
                bi_5, iva_5,   # Cols F, G
                bi_10, iva_10, # Cols H, I
                bi_21, iva_21, # Cols J, K
                f"=SUM(F{idx},H{idx},J{idx},G{idx},I{idx},K{idx})", # Col L: CHECK TOTAL (Sum of all Bases + Quotas)
                iva_texto,
                factura.get('Archivo')
            ]
            ws.append(row_data)
            
            # Estilo para filas de datos
            fill = self.zebra_fill if idx % 2 == 0 else None
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.thin_border
                if fill:
                    cell.fill = fill
                
                # Alinear importes a la derecha y formato moneda
                # Cols numéricas: Base(4), Total(5), BI5(6), IVA5(7)... hasta CHECK TOTAL(12)
                if 4 <= col <= 12:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
        
        # Añadir Fila de Totales
        ultima_fila = len(datos_facturas) + 2
        ws.cell(row=ultima_fila, column=1, value="TOTALES").font = Font(bold=True)
        
        # Sumar columnas numéricas (D a L -> 4 a 12)
        # Sumar solo columnas específicas para el resumen (D, E, L -> 4, 5, 12)
        columnas_resumen = [4, 5, 12] # Base Imponible, Total, CHECK TOTAL
        letras_columnas = {4: 'D', 5: 'E', 12: 'L'}

        for col_idx in columnas_resumen:
            let = letras_columnas[col_idx]
            formula = f"=SUM({let}2:{let}{ultima_fila-1})"
            cell = ws.cell(row=ultima_fila, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarillo para resaltar
        
        # Ajustar ancho de columnas
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 5

        # Añadir filtros
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(output_path)
